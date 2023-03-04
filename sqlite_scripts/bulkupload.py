import sqlite3
from sqlite3 import Error
import openpyxl

workbook = openpyxl.load_workbook('upload.xlsx')
sheet = workbook['NES']

def create_connection(db_file):
    """ create a database connection to the SQLite database
        specified by db_file
    :param db_file: database file
    :return: Connection object or None
    """
    conn = None
    try:
        conn = sqlite3.connect(db_file)
        return conn
    except Error as e:
        print(e)

    return conn

def create_game(conn, game):
    """
    Create a new project into the projects table
    :param conn:
    :param project:
    :return: game id
    """
    sql = ''' INSERT INTO game(game_title,platform,year_released,publisher,licensed,game_notes)
              VALUES(?,?,?,?,?,?) '''
    cur = conn.cursor()
    cur.execute(sql, game)
    conn.commit()
    print(f'{cur.lastrowid}')
    return cur.lastrowid


conn = create_connection(r"data\retro.db")

for i in range (3, 5001):
    #if column A is blank stop script
    if sheet.cell(row=i, column=2).value == None:
        break
    else:
        #define row of each interation for unit number in column A
        #idA = sheet.cell(row=i, column=1).value
        titleB = sheet.cell(row=i, column=2).value
        platformC = sheet.cell(row=i, column=3).value
        yearD = sheet.cell(row=i, column=4).value
        publisherE = sheet.cell(row=i, column=5).value
        licenseF = sheet.cell(row=i, column=6).value
        notesG = sheet.cell(row=i, column=7).value

        game=[]
        game.append(titleB)
        game.append(platformC)
        game.append(yearD)
        game.append(publisherE)
        game.append(licenseF)
        game.append(notesG)
        print(game)
        
        create_game(conn, game)
        del game

conn.close()        